import openpyxl
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.utils import get_column_letter
from rest_framework import generics
from createnews.models import NewsItem
from django.views import generic
from createnews.serializers import NewsItemSerializer


class IndexView(generic.ListView):
    template_name = 'news/index.html'
    context_object_name = 'news_list'

    def get_queryset(self):
        return NewsItem.objects.all()


def news_id(request, pk):
    context = dict(
        news_items=[NewsItem.objects.get(id=pk)]
    )
    return render(request, 'index.html', context)


def index(request):
    context = dict(
        news_items=NewsItem.objects.all()
    )
    return render(request, 'index.html', context)


class NewListJson(generics.ListAPIView):
    queryset = NewsItem.objects.all()
    serializer_class = NewsItemSerializer


class NewIDJson(generics.RetrieveAPIView):
    serializer_class = NewsItemSerializer

    def get_queryset(self):
        return NewsItem.objects.filter(id=self.kwargs['pk'])


def down_file():
    queryset = NewsItem.objects.all()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "News"
    row_num = 0
    columns = [
        ("ID", 5),
    ]

    for col_num in range(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            obj.title,
            obj.description,
            obj.image.url,
            obj.creation_date
        ]
        for col_num in range(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test.xlsx'
    wb.save(response)
    return response
